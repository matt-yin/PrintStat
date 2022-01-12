import os
from glob import glob
from datetime import datetime, timedelta
import pathlib
import shutil
import stat
import time
from openpyxl import workbook, load_workbook
for t in range(13, 0, -1):
	tempnames = []
	tempnames2 = []
	fileNames2 = []
	fileNames = []
	fileNamesPrinting = []
	printing_logFiles = []
	printing_logTempnames = []
	printing_logFullnames = []

	Path = "Z:/3. Patients for Printing/3. Completed/2021/*"
	os.chdir("Z:/3. Patients for Printing/3. Completed/2021")

	for tempnames in glob(Path):
		fileNames.append(tempnames[len(Path)-1:len(Path)+6])
		fileNames2.append(tempnames[len(Path)-1:])
		fileNames.sort()
		fileNames2.sort()

	print(len(fileNames2))
	endPos = len(fileNames2)
	startPos = endPos - 276

	#endPos = int(len(fileNames))-1

	Path = "Z:/3. Patients for Printing/2. Printing/*"
	os.chdir("Z:/3. Patients for Printing/2. Printing")

	for tempnames in glob(Path):
		fileNamesPrinting.append(tempnames[len(Path)-1:])
		fileNamesPrinting.sort()
	print(len(fileNamesPrinting))

	startPosPrinting = 0
	endPosPrinting = len(fileNamesPrinting)


	SpecificDate = datetime.now() - timedelta(days = 0)
	SpecificDate = SpecificDate.replace(hour=0, minute=0, second=0, microsecond=0)

	uniformCount = 0
	alignercoCount = 0
	alignercoRetCount = 0
	retainerCount = 0
	candidCount = 0
	smileloveCount = 0
	sequence32Count = 0
	zoomCount = 0
	usmileCount = 0
	straightsmileCount = 0
	orthofxCount = 0
	orthofxRetCount = 0
	clearForwardCount = 0
	summumCount = 0
	byteCount = 0
	surecureCount = 0


	printingPath = "Z:/3. Patients for Printing/4. Printing Log/2021/*"

	for printing_logTempnames in glob(printingPath):
		split = printing_logTempnames[49:].split()
		printing_logFiles.append(split[0])
		printing_logFullnames.append(printing_logTempnames)

	endPrintingLog = len(printing_logFiles)

	print(fileNames2[startPos])

	SpecificDate = datetime.now() - timedelta(days=t)
	print(SpecificDate.date())
	for p in range(startPos,endPos):
		# print(fileNames2[p])
		for file in os.listdir("Z:/3. Patients for Printing/3. Completed/2021/" + str(fileNames2[p]) + "/Treatment/Setup Files"):
			for x in range(0, endPrintingLog):
				printing_logFilesSplit = printing_logFiles[x].split("_")
				for x2 in range(0, len(printing_logFilesSplit)):
					if printing_logFilesSplit[x2] in fileNames2[p]:
						if printing_logFilesSplit[x2] in printing_logFullnames[x]:
							fname = pathlib.Path(printing_logFullnames[x])
							mtime = datetime.fromtimestamp(fname.stat().st_mtime)
							mtime.replace(hour=0, minute=0, second=0, microsecond=0)
							# print(mtime.day)
							# print(SpecificDate.day)
							# print()
							# print(file)
							if ".csv" not in file:
								if mtime.day == SpecificDate.day:
									if mtime.month == SpecificDate.month:
										if mtime.year == SpecificDate.year:
											# print(fileNames2[p])
											if "_UT" in fileNames2[p]:
												uniformCount+=1
											elif "_Retainer" in fileNames2[p]:
												# print(fileNames2[p])
												retainerCount+=1
											elif "_CD" in fileNames2[p]:
												candidCount +=1
											elif "_SL" in fileNames2[p]:
												smileloveCount+=1
											elif "_S32" in fileNames2[p]:
												sequence32Count+=1
											elif "_ZM" in fileNames2[p]:
												zoomCount+=1
											elif "_ACO" in fileNames2[p]:
												alignercoCount+=1
											elif "_ACO_RET" in fileNames2[p]:
												alignercoRetCount+=1
											elif "_USM" in fileNames2[p]:
												usmileCount+=1
											elif "_SSS" in fileNames2[p]:
												straightsmileCount+=1
											elif "_OFX" in fileNames2[p]:
												orthofxCount+=1
											elif "_OFX_RET" in fileNames2[p]:
												orthofxRetCount+=1
											elif "_CF" in fileNames2[p]:
												clearForwardCount+=1
											elif "_SDL" in fileNames2[p]:
												summumCount+=1
											elif "_BYT" in fileNames2[p]:
												byteCount+=1
											else:
												surecureCount+=1

	for p in range(startPosPrinting,endPosPrinting):
		# print(fileNamesPrinting[p])
		for file in os.listdir("Z:/3. Patients for Printing/2. Printing/" + str(fileNamesPrinting[p]) + "/Treatment/Setup Files"):
			for x in range(0, endPrintingLog):
				printing_logFilesSplit = printing_logFiles[x].split("_")
				for x2 in range(0, len(printing_logFilesSplit)):
					if printing_logFilesSplit[x2] in fileNamesPrinting[p]:
						if printing_logFilesSplit[x2] in printing_logFullnames[x]:
							fname = pathlib.Path(printing_logFullnames[x])
							mtime = datetime.fromtimestamp(fname.stat().st_mtime)
							mtime.replace(hour=0, minute=0, second=0, microsecond=0)
							# print(mtime.day)
							# print(SpecificDate.day)
							# print()
							if ".csv" not in file:
								if mtime.day == SpecificDate.day:
									if mtime.month == SpecificDate.month:
										if mtime.year == SpecificDate.year:
											# print(fileNamesPrinting[p])

											if "_UT" in fileNamesPrinting[p]:
												uniformCount+=1
											if "_Retainer" in fileNamesPrinting[p]:
												retainerCount+=1
											if "_CD" in fileNamesPrinting[p]:
												candidCount +=1
											if "_SL" in fileNamesPrinting[p]:
												smileloveCount+=1
											if "_S32" in fileNamesPrinting[p]:
												sequence32Count+=1
											if "_ZM" in fileNamesPrinting[p]:
												zoomCount+=1
											if "_ACO" in fileNamesPrinting[p]:
												alignercoCount+=1
											if "_ACO_RET" in fileNamesPrinting[p]:
												alignercoRetCount+=1
											if "_USM" in fileNamesPrinting[p]:
												usmileCount+=1
											if "_SSS" in fileNamesPrinting[p]:
												straightsmileCount+=1
											if "_OFX" in fileNamesPrinting[p]:
												orthofxCount+=1
											if "_OFX_RET" in fileNamesPrinting[p]:
												orthofxRetCount+=1
											if "_CF" in fileNamesPrinting[p]:
												clearForwardCount+=1
											if "_SDL" in fileNamesPrinting[p]:
												summumCount+=1
											if "_BYT" in fileNamesPrinting[p]:
												byteCount+=1
											if "_UT" not in fileNamesPrinting[p] and "_Retainer" not in fileNamesPrinting[p] and "_CD" not in fileNamesPrinting[p] and "_SL" not in fileNamesPrinting[p] and "_S32" not in fileNamesPrinting[p] and "_ZM" not in fileNamesPrinting[p] and "_ACO" not in fileNamesPrinting[p] and "_ACO_RET" not in fileNamesPrinting[p] and "_USM" not in fileNamesPrinting[p] and "_SSS" not in fileNamesPrinting[p] and "_OFX" not in fileNamesPrinting[p] and "_OFX_RET" not in fileNamesPrinting[p] and "_CF" not in fileNamesPrinting[p] and "_SDL" not in fileNamesPrinting[p] and "_BYT" not in fileNamesPrinting[p]:
												surecureCount+=1

	# print(printing_logFilesSplit)
	print("Uniform Aligners Printed: {}".format(uniformCount))
	print("AlignerCO Aligners Printed: {}".format(alignercoCount))
	print("Retainer Aligners Printed: {}".format(retainerCount))

	os.chdir("Z:/5. Shipping/8. Total number of aligners printed")
	wb = load_workbook("ProductionCountingWorksheet.xlsx")
	ws = wb.active

	uniformCount = uniformCount - 1


	row_count = ws.max_row
	ws.cell(row = row_count+1, column = 1).value = str(SpecificDate.date())
	ws.cell(row = row_count+1, column = 2).value = uniformCount
	ws.cell(row = row_count+1, column = 3).value = candidCount
	ws.cell(row = row_count+1, column = 4).value = smileloveCount
	ws.cell(row = row_count+1, column = 5).value = sequence32Count
	ws.cell(row = row_count+1, column = 6).value = zoomCount
	ws.cell(row = row_count+1, column = 7).value = alignercoCount
	ws.cell(row = row_count+1, column = 8).value = alignercoRetCount
	ws.cell(row = row_count+1, column = 9).value = usmileCount
	ws.cell(row = row_count+1, column = 10).value = straightsmileCount
	ws.cell(row = row_count+1, column = 11).value = orthofxCount
	ws.cell(row = row_count+1, column = 12).value = orthofxRetCount
	ws.cell(row = row_count+1, column = 13).value = clearForwardCount
	ws.cell(row = row_count+1, column = 14).value = summumCount
	ws.cell(row = row_count+1, column = 15).value = byteCount
	ws.cell(row = row_count+1, column = 16).value = surecureCount
	ws.cell(row = row_count+1, column = 17).value = retainerCount
	ws.cell(row = row_count+1, column = 18).value = '=SUM(B' + str(row_count+1) + ':P' + str(row_count+1) + ')'

	# print(row_count+1)

	wb.save("ProductionCountingWorksheet.xlsx")